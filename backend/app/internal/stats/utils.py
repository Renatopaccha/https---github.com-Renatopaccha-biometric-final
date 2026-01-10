import streamlit as st
import os
import io
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. GESTIÓN DE DISEÑO (CSS ROBUSTO)
# ==========================================
def load_custom_css():
    """Carga el estilo CSS usando rutas absolutas."""
    directorio_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_css = os.path.join(directorio_actual, '..', 'assets', 'style.css')
    ruta_css = os.path.abspath(ruta_css)
    
    if os.path.exists(ruta_css):
        with open(ruta_css, "r", encoding="utf-8") as f:
            st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)
    else:
        # Fallback silencioso o log
        print(f"Advertencia: No se encontró CSS en {ruta_css}")

# Alias
cargar_estilo_medico = load_custom_css

# --- HELPER DE DISEÑO: TARJETAS ---
def card_container(titulo=None):
    """Contenedor visual tipo tarjeta."""
    if titulo:
        st.markdown(f"""<h3 style="margin-bottom: 1rem;">{titulo}</h3>""", unsafe_allow_html=True)
    return st.container()


# ==========================================
# 2. MOTOR DE REPORTES EXCEL (PREMIUM)
# ==========================================
MAX_REPORT_ITEMS = 50

def agregar_al_reporte(tipo, titulo, data):
    """Añade items a la memoria de la sesión (Carrito)."""
    if 'reporte_items' not in st.session_state:
        st.session_state['reporte_items'] = []
    
    if len(st.session_state['reporte_items']) < MAX_REPORT_ITEMS:
        st.session_state['reporte_items'].append({'tipo': tipo, 'titulo': titulo, 'data': data})
        st.toast(f"✅ '{titulo}' añadido al reporte.", icon="📋")
    else:
        st.warning("⚠️ El reporte está lleno. Descárgalo para limpiar.")

def boton_guardar_grafico(fig, titulo_grafico, key_unica):
    """Botón para guardar gráficos (Matplotlib/Seaborn) en el reporte."""
    if st.button(f"➕ Añadir Gráfica", key=key_unica, help="Agrega esta imagen al Excel final"):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches='tight', dpi=150)
        buf.seek(0)
        agregar_al_reporte('img', titulo_grafico, buf)

def _excel_paper_descriptiva(df_plano, orientacion="Horizontal (como SPSS)", sheet_name="Resultados"):
    """
    Exporta un Excel con estilo académico moderno (paper/tesis),
    intentando replicar el look de la tabla mostrada en pantalla.
    df_plano debe incluir columna 'Variable' + columnas de métricas.
    """
    # --- estilos base ---
    ACCENT = "0B3A82"
    GROUP_BG = "E9EEF7"   # azul muy suave
    HEADER_BG = "F7FAFF"  # casi blanco
    ZEBRA_BG = "FAFAFA"
    LINE = "D0D7E2"

    thin = Side(style="thin", color=LINE)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_header = Font(name="Calibri", size=12, bold=True, color=ACCENT)
    font_group  = Font(name="Calibri", size=10, bold=True, color="374151")
    font_metric = Font(name="Calibri", size=11, bold=True, color="111827")
    font_cell   = Font(name="Calibri", size=11, bold=False, color="111827")

    fill_header = PatternFill("solid", fgColor=HEADER_BG)
    fill_group  = PatternFill("solid", fgColor=GROUP_BG)
    fill_zebra  = PatternFill("solid", fgColor=ZEBRA_BG)

    align_left  = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center= Alignment(horizontal="center", vertical="center")

    # --- preparar workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # --- normalizar dataframe ---
    if df_plano is None or df_plano.empty:
        ws["A1"] = "Sin datos para exportar."
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()

    if "Variable" not in df_plano.columns:
        # intenta recuperarla si viene como index
        df_plano = df_plano.copy()
        df_plano.insert(0, "Variable", df_plano.index.astype(str))

    df_indexed = df_plano.set_index("Variable")
    metrics_present = list(df_indexed.columns)
    variables = list(df_indexed.index)

    # grupos de métricas
    grupos_map = {
        "TENDENCIA CENTRAL": ["N", "Media", "Mediana", "Moda", "Suma", "M. Geom.", "IC 95%"],
        "DISPERSIÓN": ["D.E.", "Varianza", "CV %", "Mín", "Máx", "Rango", "IQR", "E.E.M."],
        "FORMA Y DIST.": ["Asimetría", "Curtosis", "P-Normalidad"],
        "PERCENTILES": [c for c in metrics_present if c.startswith("P") and c[1:].isdigit()],
    }

    ordered_groups = []
    for g, cols in grupos_map.items():
        cols_ok = [c for c in cols if c in metrics_present]
        if cols_ok:
            ordered_groups.append((g, cols_ok))

    def fmt(metric, val):
        if val is None:
            return ""
        try:
            import numpy as np
            if isinstance(val, float) and np.isnan(val):
                return ""
        except Exception:
            pass

        if metric == "N":
            try: return str(int(val))
            except Exception: return str(val)

        if metric == "P-Normalidad":
            try:
                v = float(val)
                return "<0.001" if v < 0.001 else f"{v:.3f}"
            except Exception:
                return str(val)

        if metric == "IC 95%":
            return str(val)

        try:
            return f"{float(val):.2f}"
        except Exception:
            return str(val)

    # =========================================================
    # ORIENTACIÓN VERTICAL (estadísticos hacia abajo)  ✅
    # =========================================================
    if orientacion == "Vertical (estadísticos hacia abajo)":
        # Header row
        ws.cell(row=1, column=1, value="Estadístico")
        ws.cell(row=1, column=1).font = font_header
        ws.cell(row=1, column=1).fill = fill_header
        ws.cell(row=1, column=1).alignment = align_left
        ws.cell(row=1, column=1).border = border_all

        col = 2
        for v in variables:
            c = ws.cell(row=1, column=col, value=str(v))
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_left
            c.border = border_all
            col += 1

        r = 2
        ncols = 1 + len(variables)

        for group_name, cols_in_group in ordered_groups:
            # fila de grupo (merge)
            ws.cell(row=r, column=1, value=group_name)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)

            gcell = ws.cell(row=r, column=1)
            gcell.font = font_group
            gcell.fill = fill_group
            gcell.alignment = align_left
            gcell.border = border_all

            for cc in range(2, ncols + 1):
                tmp = ws.cell(row=r, column=cc)
                tmp.fill = fill_group
                tmp.border = border_all

            r += 1

            zebra = False
            for metric in cols_in_group:
                # metric name
                c0 = ws.cell(row=r, column=1, value=str(metric))
                c0.font = font_metric
                c0.alignment = align_left
                c0.border = border_all
                if zebra:
                    c0.fill = fill_zebra

                col = 2
                for v in variables:
                    val = fmt(metric, df_indexed.loc[v, metric])
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = font_cell
                    c.alignment = align_right
                    c.border = border_all
                    if zebra:
                        c.fill = fill_zebra
                    col += 1

                zebra = not zebra
                r += 1

    # =========================================================
    # ORIENTACIÓN HORIZONTAL (como SPSS)
    # (export simple: variables en filas + métricas en columnas)
    # =========================================================
    else:
        # Header row simple
        headers = ["Variable"] + metrics_present

        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=str(h))
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_left if j == 1 else align_center
            c.border = border_all

        r = 2
        zebra = False
        for v in variables:
            # Variable name
            c0 = ws.cell(row=r, column=1, value=str(v))
            c0.font = font_metric
            c0.alignment = align_left
            c0.border = border_all
            if zebra:
                c0.fill = fill_zebra

            col = 2
            for metric in metrics_present:
                val = fmt(metric, df_indexed.loc[v, metric])
                c = ws.cell(row=r, column=col, value=val)
                c.font = font_cell
                c.alignment = align_right
                c.border = border_all
                if zebra:
                    c.fill = fill_zebra
                col += 1

            zebra = not zebra
            r += 1

    # --- ajustes finales ---
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # autosize columnas (simple)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        width = min(max(12, max_len + 2), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def boton_guardar_tabla(df, titulo_tabla, key_unica, orientacion="Horizontal (como SPSS)"):
    """Botón dual: Añadir al reporte O Descargar Excel individualmente ahora."""
    c1, c2 = st.columns([1, 1])
    
    # Botón 1: Añadir al carrito
    with c1:
        if st.button(f"➕ Añadir al Reporte", key=key_unica):
            # Generar excel con estilo para el reporte
            excel_bytes = _excel_paper_descriptiva(df, orientacion=orientacion, sheet_name=titulo_tabla)
            item = {'tipo': 'df', 'titulo': titulo_tabla, 'data': df, 'excel_bytes': excel_bytes}
            
            if 'reporte_items' not in st.session_state:
                st.session_state['reporte_items'] = []
            
            if len(st.session_state['reporte_items']) < MAX_REPORT_ITEMS:
                st.session_state['reporte_items'].append(item)
                st.toast(f"✅ '{titulo_tabla}' añadido al reporte.", icon="📋")
            else:
                st.warning("⚠️ El reporte está lleno. Descárgalo para limpiar.")
            
    # Botón 2: Descarga Inmediata (Excel Estilizado)
    with c2:
        try:
            excel_bytes = _excel_paper_descriptiva(df, orientacion=orientacion, sheet_name=titulo_tabla)
            st.download_button(
                label="⬇️ Descargar Excel",
                data=excel_bytes,
                file_name=f"{titulo_tabla}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{key_unica}"
            )
        except Exception as e:
            st.error(f"Error generando Excel: {e}")

def boton_guardar_grafico(fig, nombre_archivo: str, key: str):
    """
    Botones: Añadir al Reporte + Descargar gráfico.
    Intenta PNG (alta calidad) usando Plotly Kaleido.
    Si no está disponible, exporta como HTML interactivo.
    """
    col_a, col_b = st.columns([1, 1])

    # Intentar PNG primero
    png_bytes = None
    try:
        png_bytes = fig.to_image(format="png", scale=3)  # requiere kaleido
    except Exception:
        png_bytes = None

    html_bytes = None
    try:
        html_str = fig.to_html(full_html=True, include_plotlyjs="cdn")
        html_bytes = html_str.encode("utf-8")
    except Exception:
        html_bytes = None

    # Elegir formato disponible (prefer PNG)
    if png_bytes is not None:
        file_bytes = png_bytes
        filename = f"{nombre_archivo}.png"
        mime = "image/png"
        filetype = "png"
    else:
        file_bytes = html_bytes if html_bytes is not None else b""
        filename = f"{nombre_archivo}.html"
        mime = "text/html"
        filetype = "html"

    # Botón descargar
    with col_a:
        st.download_button(
            "⬇️ Descargar Gráfico",
            data=file_bytes,
            file_name=filename,
            mime=mime,
            key=f"dl_{key}"
        )

    # Botón añadir al reporte
    with col_b:
        if st.button("➕ Añadir al Reporte", key=f"add_{key}"):
            if 'reporte_items' not in st.session_state:
                st.session_state['reporte_items'] = []

            if len(st.session_state['reporte_items']) < MAX_REPORT_ITEMS:
                st.session_state['reporte_items'].append({
                    "tipo": "grafico",
                    "nombre": nombre_archivo,
                    "extension": filetype,
                    "bytes": file_bytes
                })
                st.toast(f"✅ '{nombre_archivo}' añadido al reporte.", icon="📊")
            else:
                st.warning("⚠️ El reporte está lleno. Descárgalo para limpiar.")



# ==========================================
# 3. UTILIDADES ADICIONALES
# ==========================================
@st.cache_data
def load_lottieurl(url: str):
    """Carga animaciones Lottie desde URL."""
    try:
        r = requests.get(url)
        return r.json() if r.status_code == 200 else None
    except:
        return None
